#!/bash/bin

import json
import os
import re
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, borders, colors
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    '''
    记录写入Excel使用的格式器
    APIs:
    get_title_map               : 获取中英文对照的格式标题行
        return {str:str}: 返回dict对象. 英: 中
    get_inheritRight_int_map    : 获取数字形式对应的 权限应用于
        return {int:str}: 返回dict对象
    get_inheritRight_mask_map   : 获取掩码形式对应的 权限应用于
        return {int:str}: 返回dict对象
    get_propagateInherit_map    : 获取数字形式对应的 传播权限
        return {int:str}: 返回dict对象
    get_parentInherit_map       : 获取数字形式对应的 从父对象继承
        return {int:str}: 返回dict对象
    '''

    @staticmethod
    def get_title_map():
        '''
            获取中英文对照的格式标题行
        '''
        return {
            'path': '查询路径',
            'domain': '所属组/域',
            'user': '用户名称',
            'fullAccessMask': '完整的权限掩码',
            'accessMask': '不包含继承关系的权限掩码',
            'inherit': '继承关系',
            'parentInherit': '从父对象继承',
            'propagateInherit': '传播继承',
            'inheritRight': '权限的应用场景',
            'isAllow': '是否允许的权限',
            'fullControl': '完全控制',
            'readData_listDir': '列出文件夹/读取数据',
            'readAttr': '读取属性',
            'readExtAttr': '读取扩展属性',
            'readPermiss': '读取权限',
            'execute_traverse': '遍历文件夹/执行文件',
            'writeData_addFile': '创建文件/写入数据',
            'appendData_addSubdir': '创建文件夹/附加数据',
            'writeAttr': '写入属性',
            'writeExtAttr': '写入扩展属性',
            'delete': '删除',
            'deleteChild': '删除子文件夹及文件',
            'changePermiss': '更改权限',
            'takeOwner': '取得所有权',
            'sync': '同步'
        }

    @staticmethod
    def get_inheritRight_int_map():
        '''
            获取数字形式对应的 权限应用于
        '''
        return {  #
            0: '只有该文件夹',
            1: '只有子文件夹',
            2: '只有文件',
            3: '此文件夹和子文件夹',
            4: '此文件夹和文件',
            5: '仅子文件夹和文件',
            6: '此文件夹、子文件夹和文件'
        }

    @staticmethod
    def get_inheritRight_mask_map():
        '''
            获取掩码形式对应的 权限应用于
        '''
        return {  #
            '': '只有该文件夹',
            '(CI)(IO)': '只有子文件夹',
            '(OI)(IO)': '只有文件',
            '(CI)': '此文件夹和子文件夹',
            '(OI)': '此文件夹和文件',
            '(OI)(CI)(IO)': '仅子文件夹和文件',
            '(OI)(CI)': '此文件夹、子文件夹和文件'
        }

    @staticmethod
    def get_propagateInherit_map():
        '''
            获取数字形式对应的 传播权限
        '''
        return {  #
            0: '不传播继承',
            1: '传播继承'
        }

    @staticmethod
    def get_parentInherit_map():
        '''
            获取数字形式对应的 从父对象继承
        '''
        return {  #
            0: '不继承',
            1: '继承'
        }


def AuthsExport(access_map_: dict, file_path_: str, file_export_: int = 0, *args, **kwargs):
    '''
        将查询结果导出到指定格式
        access_map_ : Access查询结果
        file_path_  : 写入的文件路径
        flag_export_: 导出选项.  0 不导出 | 1 导出Excel | 2 导出Json
        *args       : 不定参数集. list列表允许输入多个参数
        **kwargs    : 不定参数集. dict集合允许输入多个键值对
    
    SUCCESS
    return True     : 成功返回True
    
    ERROR
    return False    : 错误返回False
    '''
    if not access_map_ or file_export_ == 0:
        return False
    if file_export_ == 1:
        WriteExcel(access_map_, file_path_=file_path_)
    elif file_export_ == 2:
        for v in kwargs.values():
            WriteJson(access_map_, file_path_=file_path_, kwargs=v)


def WriteExcel(obj_: dict, file_path_: str, *args, **kwargs):
    '''
        输出json数据到Excel文件
        obj_        : 读取json对象
        file_path_  : 写入的文件路径
        *args       : 不定参数集. list列表允许输入多个参数
        **kwargs    : 不定参数集. dict集合允许输入多个键值对
    
    SUCCESS
    return True     : 成功返回True
    
    ERROR
    return False    : 错误返回False
    '''
    wb = Workbook()
    ws = wb.active

    # 记录已写入的行数
    _num_rows = 0
    # 写入标题行
    _num_title = 0
    for title_k, title_v in ExcelFormatter.get_title_map().items():
        _num_title += 1
        ws[get_column_letter(_num_title) + str(1)] = title_k
        ws[get_column_letter(_num_title) + str(2)] = title_v
    _num_rows += 2

    # 写入内容行
    if 'dirs' not in obj_.keys():
        # 仅查询当前路径
        for cur_k1, cur_v1 in obj_.items():
            if cur_v1['accessState'] == '拒绝访问':
                _num_rows += 1
                ws['A' + str(_num_rows)] = cur_k1
                for v in range(2, 26):
                    ws[get_column_letter(v) + str(_num_rows)] = '拒绝访问'
                continue
            for cur_v2 in cur_v1['accessState']:
                _num_rows += 1
                ws['A' + str(_num_rows)] = cur_k1
                ws['B' + str(_num_rows)] = str(cur_v2['domain'])
                ws['C' + str(_num_rows)] = str(cur_v2['user'])
                ws['D' + str(_num_rows)] = str(cur_v2['fullAccessMask'])
                ws['E' + str(_num_rows)] = '(%s)' % str.join(',', cur_v2['accessMask'])
                ws['F' + str(_num_rows)] = str(cur_v2['inherit'])
                ws['G' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_parentInherit_map().items() if str(cur_v2['parentInherit']) == str(k)][0]
                ws['H' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_propagateInherit_map().items() if str(cur_v2['propagateInherit']) == str(k)][0]
                ws['I' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_inheritRight_int_map().items() if str(cur_v2['inheritRight']) == str(k)][0]
                ws['J' + str(_num_rows)] = str(cur_v2['isAllow'])
                ws['K' + str(_num_rows)] = str(cur_v2['fullControl'])
                ws['L' + str(_num_rows)] = str(cur_v2['readData_listDir'])
                ws['M' + str(_num_rows)] = str(cur_v2['readAttr'])
                ws['N' + str(_num_rows)] = str(cur_v2['readExtAttr'])
                ws['O' + str(_num_rows)] = str(cur_v2['readPermiss'])
                ws['P' + str(_num_rows)] = str(cur_v2['execute_traverse'])
                ws['Q' + str(_num_rows)] = str(cur_v2['writeData_addFile'])
                ws['R' + str(_num_rows)] = str(cur_v2['appendData_addSubdir'])
                ws['S' + str(_num_rows)] = str(cur_v2['writeAttr'])
                ws['T' + str(_num_rows)] = str(cur_v2['writeExtAttr'])
                ws['U' + str(_num_rows)] = str(cur_v2['delete'])
                ws['V' + str(_num_rows)] = str(cur_v2['deleteChild'])
                ws['W' + str(_num_rows)] = str(cur_v2['changePermiss'])
                ws['X' + str(_num_rows)] = str(cur_v2['takeOwner'])
                ws['Y' + str(_num_rows)] = str(cur_v2['sync'])
    else:
        # 递归查询路径
        for rcur_v1 in obj_.values():
            for rcur_k2, rcur_v2 in rcur_v1.items():
                if rcur_v2['accessState'] == '拒绝访问':
                    _num_rows += 1
                    ws['A' + str(_num_rows)] = rcur_k2
                    for v in range(2, 26):
                        ws[get_column_letter(v) + str(_num_rows)] = '拒绝访问'
                    continue
                for rcur_v3 in rcur_v2['accessState']:
                    _num_rows += 1
                    ws['A' + str(_num_rows)] = rcur_k2
                    ws['B' + str(_num_rows)] = str(rcur_v3['domain'])
                    ws['C' + str(_num_rows)] = str(rcur_v3['user'])
                    ws['D' + str(_num_rows)] = str(rcur_v3['fullAccessMask'])
                    ws['E' + str(_num_rows)] = '(%s)' % str.join(',', rcur_v3['accessMask'])
                    ws['F' + str(_num_rows)] = str(rcur_v3['inherit'])
                    ws['G' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_parentInherit_map().items() if str(rcur_v3['parentInherit']) == str(k)][0]
                    ws['H' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_propagateInherit_map().items()
                                                if str(rcur_v3['propagateInherit']) == str(k)][0]
                    ws['I' + str(_num_rows)] = [v for k, v in ExcelFormatter.get_inheritRight_int_map().items() if str(rcur_v3['inheritRight']) == str(k)][0]
                    ws['J' + str(_num_rows)] = str(rcur_v3['isAllow'])
                    ws['K' + str(_num_rows)] = str(rcur_v3['fullControl'])
                    ws['L' + str(_num_rows)] = str(rcur_v3['readData_listDir'])
                    ws['M' + str(_num_rows)] = str(rcur_v3['readAttr'])
                    ws['N' + str(_num_rows)] = str(rcur_v3['readExtAttr'])
                    ws['O' + str(_num_rows)] = str(rcur_v3['readPermiss'])
                    ws['P' + str(_num_rows)] = str(rcur_v3['execute_traverse'])
                    ws['Q' + str(_num_rows)] = str(rcur_v3['writeData_addFile'])
                    ws['R' + str(_num_rows)] = str(rcur_v3['appendData_addSubdir'])
                    ws['S' + str(_num_rows)] = str(rcur_v3['writeAttr'])
                    ws['T' + str(_num_rows)] = str(rcur_v3['writeExtAttr'])
                    ws['U' + str(_num_rows)] = str(rcur_v3['delete'])
                    ws['V' + str(_num_rows)] = str(rcur_v3['deleteChild'])
                    ws['W' + str(_num_rows)] = str(rcur_v3['changePermiss'])
                    ws['X' + str(_num_rows)] = str(rcur_v3['takeOwner'])
                    ws['Y' + str(_num_rows)] = str(rcur_v3['sync'])

    # 调整Excel格式: 水平垂直居中
    for row in ws['A1:Y2']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    side = Side(
        style=borders.
        BORDER_THIN,  # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
        color=colors.BLACK,  # 边框颜色，16进制rgb表示
    )
    # 调整Excel格式: 行高列宽自适应
    max_rows_len_list = []  # 按行记录每行的最长字符串长度
    max_cols_len_list = []  # 按列记录每列的最长字符串长度
    for row in range(1, ws.max_row + 1):
        rows_value_len_list = []  # 记录当前行所有字符串
        for col in range(1, ws.max_column + 1):
            ws_cell = ws[get_column_letter(col) + str(row)]
            # 调整Excel格式: 添加黑色边框
            ws_cell.border = Border(
                top=side,  # 上
                bottom=side,  # 下
                left=side,  # 左
                right=side  # 右
            )
            if ws_cell.value:
                rows_value_len_list.append(ws_cell.value)
                if row == 1:
                    cols_value_len_list = []  # 记录当前列所有字符串
                    for rrow in range(1, ws.max_row + 1):
                        cols_value_len_list.append(ws[get_column_letter(col) + str(rrow)].value)
                    # 计算当前列所有字符串的长度
                    count_cols_value_dict = {v: (1.2 * len(re.findall(r'[\u4e00-\u9fa5]', str(v))) + len(str(v))) for v in cols_value_len_list}
                    # 计算当前列的最长字符串长度
                    max_count_cols_value = max(count_cols_value_dict.values())
                    max_cols_len_list.append(max_count_cols_value)
        # 计算当前行所有字符串的长度
        count_rows_value_dict = {v: (1.2 * len(re.findall(r'[\u4e00-\u9fa5]', str(v))) + len(str(v))) for v in rows_value_len_list}
        # 计算当前行的最长字符串长度
        max_count_rows_value = max(count_rows_value_dict.values())
        max_rows_len_list.append(max_count_rows_value)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = max_cols_len_list[col - 1] + 2

    wb.save(file_path_)
    print('\nExcel文件写入路径: %s' % file_path_)
    return True


def WriteJson(json_: dict, file_path_: str, *args, **kwargs):
    '''
        输出json数据到文件
        json_       : 写入json的内容, dict格式
        file_path_  : 写入的文件路径
        *args       : 不定参数集. list列表允许输入多个参数
        **kwargs    : 不定参数集. dict集合允许输入多个键值对
    
    SUCCESS
    return True     : 成功返回True
    
    ERROR
    return False    : 错误返回False
    '''
    writeJson = {}
    for v in kwargs.values():
        for k1, v1 in v.items():
            writeJson[k1] = v1
    writeJson['writeTime'] = time.strftime("%Y-%m-%d %H:%M", time.localtime(time.time()))
    writeJson['result'] = json_
    with open(file_path_, 'w', encoding='utf-8') as f:
        f.write(json.dumps(writeJson, indent=2, ensure_ascii=False))
    print('\nJson文件写入路径: %s' % file_path_)
    return True


def depth_walk(paths_: str, depthLevel_: int = 2):
    walks = []
    walk_parse(walks, paths_, depthLevel_)
    return walks


def walk_parse(walks_: list, paths_: str, depthLevel_: int = 2, topDepth_: int = -1):
    '''
        使用os.walk递归遍历指定depthlevel的文件夹
        paths_      : 文件夹的绝对路径 (必填)
        depthLevel_ : 递归遍历的深度, 默认递归到第二级 (必填)

    SUCCESS
    Write walks_    : 成功则写入walks_
    
    ERROR
    return          : 错误返回
    '''
    if '$RECYCLE.BIN' in paths_:
        return
    if paths_[-1] != '\\':
        paths_ += '\\'
    topDepth_ = paths_.count('\\') if topDepth_ == -1 else topDepth_
    for (root, dirs, files) in os.walk(paths_):
        walks_.append({'root': root, 'dirs': dirs, 'files': files})
        depth = root.count('\\') if root[-1] == '\\' else root.count('\\') + 1
        end_depth = topDepth_ + depthLevel_ - 1
        if depth < end_depth:
            dirs = [os.path.join(root, v) for v in dirs]
            for dir in dirs:
                if os.path.isdir(dir):
                    walk_parse(walks_, dir, depthLevel_, topDepth_)
        else:
            return
        return


if __name__ == '__main__':
    pass
